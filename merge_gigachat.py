import pandas as pd
import openpyxl

def run():
    INPUT_FILE1 = "source_expanded.xlsx"
    INPUT_FILE2 = "Result_GigaChat.xlsx"
    OUTPUT_FILE = "merged_gigachat_finalfile.xlsx"

    df1 = pd.read_excel(INPUT_FILE1)
    df2 = pd.read_excel(INPUT_FILE2)

    new_df = pd.concat([
        df1.iloc[:, [0]],       # 1 колонка из df1
        df2.iloc[:, [2]],       # 2 колонка из df2
        df2.iloc[:, [3]],       
        df2.iloc[:, [4]],     
        df1.iloc[:, [1]],    
        df2.iloc[:, [0]],       
        df1.iloc[:, [2]],      
        df2.iloc[:, [1]],     
        df1.iloc[:, [3]],      
        df2.iloc[:, [5]],
        df1.iloc[:, [4]],
        df2.iloc[:, [6]],
        df1.iloc[:, [5]],
        df2.iloc[:, [7]],
        df2.iloc[:, [8]],
        df1.iloc[:, [6]],
        df2.iloc[:, [9]]
        

    ], axis=1)

    new_df.to_excel(OUTPUT_FILE, index=False)

    # подсчет для оформлениия
    col = df2.iloc[:, 0]

    counts = []
    i = 0
    while i < len(col):
        count = 1
        while i + count < len(col) and col[i + count] == col[i]:
            count += 1
        counts.append(count)
        i += count

    # openpyxl
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    ws = wb.active

    # ширина
    for col in ws.columns:
        col_letter = col[0].column_letter  # буква колонки (A, B, C...)
        ws.column_dimensions[col_letter].width = 100
        ws.column_dimensions["A"].width = 50

    # высота
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 50

    # настройка границ
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    # цвета
    colors = ["84B082", "F7C1BB"] 
    header_color = "F2FF49"
    color_index = 0 
    row_start = 2 

    ws.cell(row=1, column=1).fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    ws.cell(row=1, column=5).fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    ws.cell(row=1, column=7).fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    ws.cell(row=1, column=9).fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    ws.cell(row=1, column=11).fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    ws.cell(row=1, column=13).fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    ws.cell(row=1, column=16).fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")

    for c in counts:
        for r in range(row_start, row_start + c):
            for cell in ws[r]:
                cell.fill = openpyxl.styles.PatternFill(start_color=colors[color_index],
                                        end_color=colors[color_index],
                                        fill_type="solid")
                cell.border = thin_border
        row_start += c
        color_index = 1 - color_index  # чередуем два цвета

    # сохранение изменений
    wb.save(OUTPUT_FILE)

    print("merge.py выполнен")

if __name__ == "__main__":
    run()